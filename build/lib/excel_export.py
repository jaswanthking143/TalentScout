"""
excel_export.py
Builds the TalentScout analysis output as a styled .xlsx workbook instead
of a Tkinter popup: a cover sheet (project name/description + run summary)
and a ranked results sheet limited to the top N candidates the HR user asked
for (e.g. Top 5 / Top 10).
"""
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import Candidate

PROJECT_NAME = "TalentScout"
PROJECT_TAGLINE = "Resume Intake & Role-Fit Analyzer"
PROJECT_DESCRIPTION = (
    "TalentScout is a desktop tool that helps HR teams cut through a stack "
    "of resumes quickly. Upload one or more PDF resumes, and the app "
    "extracts each candidate's name, contact details, skills, and "
    "experience automatically.\n\n"
    "Type in the role you're hiring for and TalentScout scores every "
    "candidate against that role's expected skill set, ranks them from "
    "best to worst fit, and highlights exactly which required skills each "
    "candidate has and which ones they're missing.\n\n"
    "This report contains the results of that analysis, limited to the "
    "top-ranked candidates you selected."
)

NAVY = "0A192F"
ACCENT = "64FFDA"
PANEL = "112240"
WHITE = "FFFFFF"
MUTED = "8892B0"


def _style_cover_sheet(ws, role: str, top_n: int, total_candidates: int):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 90

    title_fill = PatternFill("solid", fgColor=NAVY)
    for row in range(1, 4):
        for col in range(1, 3):
            ws.cell(row=row, column=col).fill = title_fill

    ws.merge_cells("B2:B2")
    c = ws["B2"]
    c.value = PROJECT_NAME
    c.font = Font(name="Calibri", size=26, bold=True, color=ACCENT)
    c.alignment = Alignment(vertical="center")

    ws["B3"] = PROJECT_TAGLINE
    ws["B3"].font = Font(name="Calibri", size=12, italic=True, color=WHITE)

    row = 5
    for para in PROJECT_DESCRIPTION.split("\n\n"):
        cell = ws.cell(row=row, column=2, value=para)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Calibri", size=11)
        ws.row_dimensions[row].height = 55
        row += 2

    row += 1
    ws.cell(row=row, column=2, value="Run Summary").font = Font(
        name="Calibri", size=13, bold=True, color=NAVY
    )
    row += 1
    summary_lines = [
        f"Role analyzed: {role.title()}",
        f"Candidates loaded: {total_candidates}",
        f"Showing top: {top_n}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for line in summary_lines:
        cell = ws.cell(row=row, column=2, value=line)
        cell.font = Font(name="Calibri", size=11)
        row += 1


def _style_results_sheet(ws, role: str, ranked: List[Candidate]):
    headers = [
        "Rank", "Name", "Email", "Phone", "Experience",
        "Match Score (%)", "Matched Skills", "Missing Skills",
    ]
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
    thin = Side(style="thin", color="1D3A63")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"

    best_score = ranked[0].match_score if ranked else 0
    for i, c in enumerate(ranked, start=1):
        row = [
            i,
            c.name,
            c.email,
            c.phone,
            c.experience,
            c.match_score,
            ", ".join(c.matched_skills) or "None",
            ", ".join(c.missing_skills) or "None",
        ]
        ws.append(row)
        r = ws.max_row
        is_best = c.match_score == best_score and best_score > 0
        row_fill = PatternFill("solid", fgColor="E9FBF6" if is_best else ("F4F6FB" if r % 2 == 0 else "FFFFFF"))
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx in (7, 8))
            if col_idx == 1 and is_best:
                cell.value = f"🏆 {i}"
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="top")

    widths = [7, 20, 26, 16, 18, 16, 40, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def export_candidates_to_excel(
    output_path: str,
    role: str,
    ranked: List[Candidate],
    top_n: int,
    total_candidates: int,
) -> str:
    """Writes a two-sheet workbook (cover + top-N ranked results) to
    output_path and returns that path."""
    top_ranked = ranked[:top_n]

    wb = Workbook()
    cover = wb.active
    cover.title = "TalentScout"
    _style_cover_sheet(cover, role, len(top_ranked), total_candidates)

    results = wb.create_sheet("Results")
    _style_results_sheet(results, role, top_ranked)

    wb.save(output_path)
    return output_path